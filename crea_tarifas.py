"""
ACTUALIZAR TARIFAS TURACO
=========================
Genera TARIFA NACIONAL COMPLETA y TARIFA INTER COMPLETA
a partir de los ficheros de Sistema (TARIFA TURACO y TARIFA TURACO INTER)
y la calculadora (CALCULADORA_TARIFA_NACIONAL.xlsx).

Uso:
    python actualizar_tarifas.py \
        --nacional  TARIFA_TURACO_-_ABR_2026.xlsx \
        --inter     TARIFA_TURACO_INTER_-_ABR2026.xlsx \
        --calc      CALCULADORA_TARIFA_NACIONAL.xlsx \
        [--margen   0.10]   # Margen de seguridad, por defecto 10%

Salida:
    TARIFA_NACIONAL_COMPLETA_<fecha>.xlsx
    TARIFA_TURACO_INTER_COMPLETA_<fecha>.xlsx
"""

import argparse
import math
import re
import sys
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def parse_euro(value):
    """Convierte '109,00 €' o '109.00' o 109.0 → float. Devuelve None si no es vendible."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s.lower() in ('no vendible', 'no_vendible', ''):
        return None
    s = s.replace('€', '').replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def pvp_min(cost_log):
    """=REDONDEAR.MAS(cost_log; 0) - 0,1"""
    return math.ceil(cost_log) - 0.1


def pvp_pub(pvp_min_val, pvpr):
    """=SI(pvp_min > pvpr; pvp_min; pvpr)"""
    return pvp_min_val if pvp_min_val > pvpr else pvpr


def rentabilidad(neto, porte, pvp_sin_iva, comision_eur):
    denom = pvp_sin_iva - comision_eur
    if denom == 0:
        return None
    return 1 - (neto + porte) / denom


def header_style(ws, row, cols, bg='1F4E79', fg='FFFFFF', bold=True):
    fill = PatternFill('solid', start_color=bg)
    font = Font(bold=bold, color=fg, name='Arial', size=10)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def num_cell(ws, row, col, value, fmt='#,##0.00'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = fmt
    return cell


def pct_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = '0.00%'
    return cell


def bool_cell(ws, row, col, value):
    ws.cell(row=row, column=col, value='SI' if value else 'NO')


# ─────────────────────────────────────────────────────────────────────────────
# CARGA DE DATOS
# ─────────────────────────────────────────────────────────────────────────────

def load_product_info(calc_path):
    """Carga Product_info de la calculadora → dict EAN → {ref, titulo, familia, subfamilia, comisiones}"""
    wb = load_workbook(calc_path, read_only=True, data_only=True)
    ws = wb['Product_info']
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]

    col = {h: i for i, h in enumerate(headers)}
    products = {}  # EAN → dict

    for row in rows[1:]:
        if not row[col.get('EAN', 1)]:
            continue
        try:
            ean = str(int(float(str(row[col['EAN']])))) if row[col['EAN']] else None
        except (ValueError, TypeError):
            continue
        if not ean:
            continue
        products[ean] = {
            'ref':        row[col['REFERENCIA']],
            'titulo':     row[col.get('TÍTULO DE PRODUCTO', col.get('TITULO DE PRODUCTO', 2))],
            'familia':    row[col.get('FAMILIA', 4)],
            'subfamilia': row[col.get('SUBFAMILIA', 5)],
            'com_amz':    row[col.get('COM_AZM', 7)],
            'com_mir':    row[col.get('COM_MIR', 8)],
            'com_mm':     row[col.get('COM_MM', 9)],
            'com_priv':   row[col.get('COM_PRIV', 10)],
            'com_c4':     row[col.get('COM_C4', 11)],
        }
    wb.close()
    return products


def load_tarifa_nacional(path):
    """Carga tarifa nacional → dict EAN → {pvpr, neto, portes}"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Header en fila 2
    headers = None
    tarifa = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if i == 2:
            headers = [str(h).strip() if h else '' for h in row]
            col = {h: j for j, h in enumerate(headers)}
            continue
        ean_raw = row[col.get('EAN', 5)]
        if not ean_raw:
            continue
        ean = str(int(float(ean_raw))) if ean_raw else None
        pvpr  = row[col.get('PVPR ', col.get('PVPR', 10))]
        neto  = row[col.get('NETO', 11)]
        porte = row[col.get('PORTES', col.get('PORTE', 12))]
        if pvpr is None or neto is None:
            continue
        tarifa[ean] = {
            'pvpr':  float(pvpr),
            'neto':  float(neto),
            'porte': float(porte) if porte else 0.0,
        }
    wb.close()
    return tarifa


def load_tarifa_inter_sheet(wb, sheet_name, col_pvp, col_porte_es, col_neto_es,
                             col_porte_local=None, col_neto_local=None,
                             header_row=2, data_start=3):
    """Carga una pestaña del INTER → list of dicts con nombre, ref_str, ean, pvp, porte_es, neto_es, porte_loc, neto_loc"""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    # Col indices are already given as positional (0-based)
    records = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row[0]:
            continue
        ean_raw = row[4]
        ean = str(int(float(str(ean_raw).replace(',', '.').replace(' ', '')))) if ean_raw else None

        pvp = parse_euro(row[col_pvp])
        porte_es = parse_euro(row[col_porte_es])
        neto_es  = parse_euro(row[col_neto_es])
        porte_loc = parse_euro(row[col_porte_local]) if col_porte_local is not None else None
        neto_loc  = parse_euro(row[col_neto_local])  if col_neto_local  is not None else None

        records.append({
            'nombre':    row[0],
            'tipo':      row[2],
            'ref_str':   str(row[3]) if row[3] else '',
            'ean':       ean,
            'pvp':       pvp,
            'porte_es':  porte_es,
            'neto_es':   neto_es,
            'porte_loc': porte_loc,
            'neto_loc':  neto_loc,
        })
    return records


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN TARIFA NACIONAL COMPLETA
# ─────────────────────────────────────────────────────────────────────────────

MARKETPLACES = [
    ('T_AMZ',  '(%) COM_AMZ', 'com_amz',  0.1815, 'AMAZON',     1.21),
    ('T_MIR',  '(%) COM_MIR', 'com_mir',  0.15,   'MIRAVIA',    1.21),
    ('T_C4',   '(%) COM_C4',  'com_c4',   0.10,   'CARREFOUR',  1.21),
    ('T_MM',   '(%) COM_MM',  'com_mm',   0.14956,'MEDIAMARKT', 1.21),
    ('T_PRIV', '(%) COM_PRIV','com_priv', 0.1452, 'PRIVALIA',   1.21),
]

NAC_HEADERS = [
    'REFERENCIA', 'EAN', 'NOMBRE COMPLETO', 'FAMILIA', 'SUBFAMILIA',
    'PVPR ', 'NETO', 'PORTES', 'MARGEN (%)', 'MARGEN (€)',
    'COST_log', 'COM (%)', 'COM (€)', 'IVA (21%)',
    'PVP MIN.', 'PVP PUB.', 'PVP SIN IVA', 'COMISION', 'RENTABILIDAD', 'DESPOSICIONADO'
]

NAC_HEADER_BG = {
    'T_AMZ':  'FF6600',
    'T_MIR':  '0070C0',
    'T_C4':   'C00000',
    'T_MM':   '00B050',
    'T_PRIV': '7030A0',
}


def build_nacional(wb_out, tarifa, products, margen_default=0.10):
    for sheet_id, com_sheet_id, com_key, com_default, mp_name, iva in MARKETPLACES:
        ws = wb_out.create_sheet(sheet_id)
        bg = NAC_HEADER_BG.get(sheet_id, '1F4E79')

        # Header row
        for j, h in enumerate(NAC_HEADERS, start=1):
            cell = ws.cell(row=1, column=j, value=h)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center')

        # Also create companion COM% sheet (lookup table)
        ws_com = wb_out.create_sheet(com_sheet_id)
        ws_com.cell(row=1, column=1, value='REFERENCIA').font = Font(bold=True)
        ws_com.cell(row=1, column=2, value='COMISION (%)').font = Font(bold=True)

        com_row = 2
        row = 2
        for ean, prod in products.items():
            if ean not in tarifa:
                continue
            tar = tarifa[ean]
            ref     = prod['ref']
            pvpr    = tar['pvpr']
            neto    = tar['neto']
            porte   = tar['porte']
            margen  = margen_default
            com_pct = prod.get(com_key) or com_default

            margen_eur  = neto / (1 - margen)
            cost_log    = margen_eur + porte
            com_eur     = cost_log / (1 - com_pct)
            iva_eur     = com_eur * iva
            pmin        = pvp_min(iva_eur)
            ppub        = pvp_pub(pmin, pvpr)
            pvp_sin_iva = ppub / iva
            comision    = pvp_sin_iva * com_pct
            rent        = rentabilidad(neto, porte, pvp_sin_iva, comision)
            despos      = ppub > pvpr

            ws.cell(row=row, column=1,  value=ref)
            ws.cell(row=row, column=2,  value=int(ean) if ean.isdigit() else ean)
            ws.cell(row=row, column=3,  value=prod['titulo'])
            ws.cell(row=row, column=4,  value=prod['familia'])
            ws.cell(row=row, column=5,  value=prod['subfamilia'])
            num_cell(ws, row, 6,  pvpr)
            num_cell(ws, row, 7,  neto)
            num_cell(ws, row, 8,  porte)
            pct_cell(ws, row, 9,  margen)
            num_cell(ws, row, 10, margen_eur)
            num_cell(ws, row, 11, cost_log)
            pct_cell(ws, row, 12, com_pct)
            num_cell(ws, row, 13, com_eur)
            num_cell(ws, row, 14, iva_eur)
            num_cell(ws, row, 15, pmin)
            num_cell(ws, row, 16, ppub)
            num_cell(ws, row, 17, pvp_sin_iva)
            num_cell(ws, row, 18, comision)
            pct_cell(ws, row, 19, rent)
            ws.cell(row=row, column=20, value='SI' if despos else 'NO')

            # COM% lookup sheet
            ws_com.cell(row=com_row, column=1, value=ref)
            ws_com.cell(row=com_row, column=2, value=com_pct)
            com_row += 1
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['C'].width = 40
        for col_letter in ['D', 'E']:
            ws.column_dimensions[col_letter].width = 22
        for col_letter in ['F', 'G', 'H', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
            ws.column_dimensions[col_letter].width = 14

    print(f"  Nacional: {row-2} productos por marketplace")


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN TARIFA INTER COMPLETA
# ─────────────────────────────────────────────────────────────────────────────

INTER_COM_DEFAULT = 0.1815
IVA_BY_COUNTRY = {
    'FR': 1.20, 'IT': 1.22, 'DE': 1.19,
    'PT': 1.23, 'NL': 1.21, 'BE': 1.21,
    'PL': 1.23, 'SE': 1.25,  # nota: se usa como divisor según completa: /1.23, /1.25
}
MARGEN_INTER = 0.05   # default inter (según fórmulas observadas en completa)
MARGEN_INTER_BIG = 0.10  # para productos no pequeños


def calc_inter_row(pvp, neto, porte, iva_factor, margen, com_pct=INTER_COM_DEFAULT):
    if pvp is None or neto is None or porte is None:
        return None
    margen_eur = neto / (1 - margen)
    cost_log   = margen_eur + porte
    comision   = cost_log / (1 - com_pct)
    # IVA: para PL/SE el factor se divide (precios ya sin IVA claro), para el resto se multiplica
    iva_val    = comision * iva_factor
    pmin       = pvp_min(iva_val)
    ppub       = pvp_pub(pmin, pvp)
    return {
        'margen':  margen_eur,
        'cost':    cost_log,
        'comis':   comision,
        'iva':     iva_val,
        'pmin':    pmin,
        'ppub':    ppub,
    }


INTER_SHEET_HEADERS = {
    'ES-XX': ['NOMBRE COMPLETO', 'Tipo Tarifa', 'REFERENCIA', 'EAN',
              'PVP', 'NETO ES', 'PORTE ES', 'MARGEN', 'COST_log',
              'COMISION', 'IVA', 'PVP MIN.', 'PVP PUB.'],
    'XX-XX': ['NOMBRE COMPLETO', 'Tipo Tarifa', 'REFERENCIA', 'EAN',
              'PVP', 'NETO LOC', 'PORTE LOC', 'MARGEN', 'COST_log',
              'COMISION', 'IVA', 'PVP MIN.', 'PVP PUB.', 'SKU LOCAL'],
    'PL_SE': ['NOMBRE COMPLETO', 'Tipo Tarifa', 'REFERENCIA', 'EAN',
              'PVP', 'NETO', 'PORTE', 'MARGEN', 'COST_log',
              'COMISION', 'IVA', 'PVP MIN.', 'PVP PUB.', 'PVP PUB. (MONEDA LOCAL)'],
}

INTER_BG = {
    'ES-FR': 'C00000', 'FR-FR': 'FF0000',
    'ES-IT': '00B050', 'IT-IT': '008000',
    'ES-DE': '0070C0', 'DE-DE': '0000FF',
    'PT':    'FFC000', 'NL':    '7030A0',
    'BE':    '00B0F0', 'PL':    'FF6600', 'SE':    '4BACC6',
}

CONV_PL = 5   # 1€ = 5 PLN
CONV_SE = 11  # 1€ = 11 SEK


def write_inter_sheet(ws, records, products, ean_map, sheet_key,
                      country_code, iva_factor, margen,
                      local_prefix=None, conv_factor=None):
    """Escribe datos calculados en una hoja del inter completa."""
    bg = INTER_BG.get(sheet_key, '1F4E79')
    hdrs = INTER_SHEET_HEADERS.get(
        'PL_SE' if conv_factor else ('XX-XX' if local_prefix else 'ES-XX'),
        INTER_SHEET_HEADERS['ES-XX']
    )

    for j, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for rec in records:
        ean = rec['ean']
        if local_prefix:
            neto  = rec['neto_loc']
            porte = rec['porte_loc']
        else:
            neto  = rec['neto_es']
            porte = rec['porte_es']

        pvp = rec['pvp']
        # Skip non-vendible
        if pvp is None or neto is None or porte is None:
            continue

        # Determine MARGEN: use 5% for smaller products (heuristic: if neto<30 use 5%, else 10%)
        m = MARGEN_INTER if neto < 30 else MARGEN_INTER_BIG

        calc = calc_inter_row(pvp, neto, porte, iva_factor, m)
        if not calc:
            continue

        # Numeric REFERENCIA from Product_info via EAN
        prod = ean_map.get(ean, {})
        ref_num = prod.get('ref', rec['ref_str'])

        ws.cell(row=row, column=1, value=rec['nombre'])
        ws.cell(row=row, column=2, value=rec['tipo'])
        ws.cell(row=row, column=3, value=ref_num)
        ws.cell(row=row, column=4, value=int(ean) if ean and ean.isdigit() else ean)
        num_cell(ws, row, 5,  pvp)
        num_cell(ws, row, 6,  neto)
        num_cell(ws, row, 7,  porte)
        num_cell(ws, row, 8,  calc['margen'])
        num_cell(ws, row, 9,  calc['cost'])
        num_cell(ws, row, 10, calc['comis'])
        num_cell(ws, row, 11, calc['iva'])
        num_cell(ws, row, 12, calc['pmin'])
        num_cell(ws, row, 13, calc['ppub'])

        if local_prefix and ref_num:
            # SKU local: FR00112, IT00112, DE00112
            ref_str = str(ref_num).replace(' ', '')
            sku_local = f"{local_prefix}{ref_str.zfill(5)}"
            ws.cell(row=row, column=14, value=sku_local)

        if conv_factor:
            # PVP en moneda local
            ws.cell(row=row, column=14, value=round(calc['ppub'] * conv_factor, 1))

        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    return row - 2


def build_inter(wb_out, inter_path, products, margen_default=0.10):
    wb_inter = load_workbook(inter_path, read_only=True, data_only=True)
    ean_map = products  # keyed by EAN string

    # ── Helper: load one Zona sheet ──────────────────────────────────────────
    def load_zona(sheet_name, col_pvp=5, col_porte=6, col_neto=7):
        ws = wb_inter[sheet_name]
        recs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 8 or not row[0]:
                continue
            ean_raw = row[4]
            ean = None
            if ean_raw:
                try:
                    ean = str(int(float(str(ean_raw).replace(',', '.').replace(' ', ''))))
                except:
                    ean = str(ean_raw)
            recs.append({
                'nombre':    row[0],
                'tipo':      row[2],
                'ref_str':   str(row[3]) if row[3] else '',
                'ean':       ean,
                'pvp':       parse_euro(row[col_pvp]),
                'porte_es':  parse_euro(row[col_porte]),
                'neto_es':   parse_euro(row[col_neto]),
                'porte_loc': None,
                'neto_loc':  None,
            })
        return recs

    # ── FRANCIA ─────────────────────────────────────────────────────────────
    fr_recs = load_tarifa_inter_sheet(
        wb_inter, 'Francia',
        col_pvp=5, col_porte_es=6, col_neto_es=7,
        col_porte_local=8, col_neto_local=9
    )
    ws_esfr = wb_out.create_sheet('ES-FR')
    n = write_inter_sheet(ws_esfr, fr_recs, products, ean_map, 'ES-FR', 'FR', 1.20, MARGEN_INTER)
    print(f"  ES-FR: {n} filas")

    ws_frfr = wb_out.create_sheet('FR-FR')
    n = write_inter_sheet(ws_frfr, fr_recs, products, ean_map, 'FR-FR', 'FR', 1.20, MARGEN_INTER, local_prefix='FR')
    print(f"  FR-FR: {n} filas")

    # ── ITALIA ──────────────────────────────────────────────────────────────
    it_recs = load_tarifa_inter_sheet(
        wb_inter, 'Italia',
        col_pvp=5, col_porte_es=6, col_neto_es=7,
        col_porte_local=8, col_neto_local=9
    )
    ws_esit = wb_out.create_sheet('ES-IT')
    n = write_inter_sheet(ws_esit, it_recs, products, ean_map, 'ES-IT', 'IT', 1.22, MARGEN_INTER)
    print(f"  ES-IT: {n} filas")

    ws_itit = wb_out.create_sheet('IT-IT')
    n = write_inter_sheet(ws_itit, it_recs, products, ean_map, 'IT-IT', 'IT', 1.22, MARGEN_INTER, local_prefix='IT')
    print(f"  IT-IT: {n} filas")

    # ── ALEMANIA ─────────────────────────────────────────────────────────────
    de_recs = load_tarifa_inter_sheet(
        wb_inter, 'Alemania',
        col_pvp=5, col_porte_es=6, col_neto_es=7,
        col_porte_local=8, col_neto_local=9
    )
    ws_esde = wb_out.create_sheet('ES-DE')
    n = write_inter_sheet(ws_esde, de_recs, products, ean_map, 'ES-DE', 'DE', 1.19, MARGEN_INTER)
    print(f"  ES-DE: {n} filas")

    ws_dede = wb_out.create_sheet('DE-DE')
    n = write_inter_sheet(ws_dede, de_recs, products, ean_map, 'DE-DE', 'DE', 1.19, MARGEN_INTER, local_prefix='DE')
    print(f"  DE-DE: {n} filas")

    # ── PORTUGAL ─────────────────────────────────────────────────────────────
    pt_recs = load_zona('Portugal', col_pvp=5, col_porte=6, col_neto=7)
    ws_pt = wb_out.create_sheet('PT')
    n = write_inter_sheet(ws_pt, pt_recs, products, ean_map, 'PT', 'PT', 1.23, MARGEN_INTER_BIG)
    print(f"  PT: {n} filas")

    # ── HOLANDA / BÉLGICA (Zona 3) ───────────────────────────────────────────
    z3_recs = load_zona('Zona 3', col_pvp=5, col_porte=6, col_neto=7)
    ws_nl = wb_out.create_sheet('NL')
    n = write_inter_sheet(ws_nl, z3_recs, products, ean_map, 'NL', 'NL', 1.21, MARGEN_INTER_BIG)
    print(f"  NL: {n} filas")

    ws_be = wb_out.create_sheet('BE')
    n = write_inter_sheet(ws_be, z3_recs, products, ean_map, 'BE', 'BE', 1.21, MARGEN_INTER_BIG)
    print(f"  BE: {n} filas")

    # ── POLONIA (Zona 4) ─────────────────────────────────────────────────────
    z4_recs = load_zona('Zona 4', col_pvp=5, col_porte=6, col_neto=7)
    ws_pl = wb_out.create_sheet('PL')
    n = write_inter_sheet(ws_pl, z4_recs, products, ean_map, 'PL', 'PL', 1.23, MARGEN_INTER_BIG, conv_factor=CONV_PL)
    print(f"  PL: {n} filas")

    # ── SUECIA (Zona 5) ──────────────────────────────────────────────────────
    z5_recs = load_zona('Zona 5', col_pvp=5, col_porte=6, col_neto=7)
    ws_se = wb_out.create_sheet('SE')
    n = write_inter_sheet(ws_se, z5_recs, products, ean_map, 'SE', 'SE', 1.25, MARGEN_INTER_BIG, conv_factor=CONV_SE)
    print(f"  SE: {n} filas")

    wb_inter.close()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Genera tarifas Turaco completas')
    parser.add_argument('--nacional', required=True, help='TARIFA_TURACO_(MES).xlsx')
    parser.add_argument('--inter',    required=True, help='TARIFA_TURACO_INTER_(MES).xlsx')
    parser.add_argument('--calc',     required=True, help='CALCULADORA_TARIFA_NACIONAL.xlsx')
    parser.add_argument('--margen',   type=float, default=0.10, help='Margen (default 0.10 = 10%%)')
    args = parser.parse_args()

    fecha = date.today().strftime('%d_%m_%Y')

    print("Cargando Product_info de calculadora...")
    products = load_product_info(args.calc)
    print(f"  {len(products)} productos en calculadora")

    print("Cargando tarifa nacional de Sistema...")
    tarifa = load_tarifa_nacional(args.nacional)
    print(f"  {len(tarifa)} líneas en tarifa nacional")

    # ── NACIONAL COMPLETA ────────────────────────────────────────────────────
    print("\nGenerando TARIFA NACIONAL COMPLETA...")
    wb_nac = Workbook()
    wb_nac.remove(wb_nac.active)  # quitar hoja vacía
    build_nacional(wb_nac, tarifa, products, margen_default=args.margen)
    out_nac = f'TARIFA_NACIONAL_COMPLETA_{fecha}.xlsx'
    wb_nac.save(out_nac)
    print(f"  → Guardado: {out_nac}")

    # ── INTER COMPLETA ───────────────────────────────────────────────────────
    print("\nGenerando TARIFA TURACO INTER COMPLETA...")
    wb_int_out = Workbook()
    wb_int_out.remove(wb_int_out.active)
    build_inter(wb_int_out, args.inter, products, margen_default=args.margen)
    out_int = f'TARIFA_TURACO_INTER_COMPLETA_{fecha}.xlsx'
    wb_int_out.save(out_int)
    print(f"  → Guardado: {out_int}")

    print("\n✓ Proceso completado.")


if __name__ == '__main__':
    main()
